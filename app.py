import streamlit as st
import pandas as pd
import requests
import time
import os
from builtins import ValueError  # 确保ValueError可用
from pandasai import SmartDatalake
from pandasai.llm.base import LLM

# --------------------------------------------------------------------------
# 1. 自定义 DeepSeek LLM 类
# --------------------------------------------------------------------------
class DeepSeekLLM(LLM):
    """优化的DeepSeek LLM集成类"""
    def __init__(self, api_key: str, model: str = "deepseek-chat", temperature: float = 0.3):
        super().__init__()
        if not api_key or not isinstance(api_key, str):
            from builtins import ValueError
            raise ValueError("无效的API密钥格式，请检查。")
        self.api_key = api_key
        self.model = model
        self.temperature = temperature
        self.api_base = "https://api.deepseek.com/v1"
        self.max_retries = 3
        self._timeout = 60  # 内部超时变量
        self.last_response = None

    @property
    def timeout(self):
        return self._timeout

    @timeout.setter
    def timeout(self, value):
        self._timeout = value

    def call(self, prompt: str, *args, **kwargs) -> str:
        """兼容父类LLM的call方法签名，调用DeepSeek API"""
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": self.model,
            "messages": [{"role": "user", "content": str(prompt)}],
            "temperature": self.temperature,
            "max_tokens": 4096 # 增加max_tokens以获得更完整的代码或分析
        }
        
        for attempt in range(self.max_retries):
            try:
                response = requests.post(
                    f"{self.api_base}/chat/completions",
                    headers=headers,
                    json=payload,
                    timeout=self.timeout
                )
                response.raise_for_status() # 如果HTTP状态码是4xx或5xx，则抛出异常
                self.last_response = response.json()
                return self.last_response["choices"][0]["message"]["content"]
            
            except requests.exceptions.HTTPError as e:
                 # 特别处理402付款错误
                if e.response.status_code == 402:
                    raise Exception("API请求失败: 402 Client Error. 付款失败或账户余额不足。请检查您的DeepSeek账户。")
                # 其他HTTP错误
                if attempt == self.max_retries - 1:
                    raise Exception(f"API请求失败: {str(e)}")
                time.sleep(2 ** attempt)

            except requests.exceptions.RequestException as e:
                if attempt == self.max_retries - 1:
                    raise Exception(f"网络或API请求失败: {str(e)}")
                time.sleep(2 ** attempt)
    
    @property
    def type(self) -> str:
        return "deepseek-llm"

# --------------------------------------------------------------------------
# 2. Streamlit 应用配置
# --------------------------------------------------------------------------
st.set_page_config(
    page_title="🧠 智能数据分析专家",
    page_icon="📊",
    layout="wide"
)

# --------------------------------------------------------------------------
# 3. 模板管理功能
# --------------------------------------------------------------------------
def manage_templates():
    """管理Excel模板文件"""
    if not os.path.exists("templates"):
        os.makedirs("templates")
    
    with st.sidebar:
        st.header("📁 模板管理")
        
        # 上传新模板
        new_template = st.file_uploader(
            "上传新模板", 
            type=["xlsx"],
            key="template_uploader"
        )
        
        if new_template:
            template_name = st.text_input("模板名称", value=new_template.name.split(".")[0])
            if st.button("保存模板"):
                template_path = os.path.join("templates", f"{template_name}.xlsx")
                with open(template_path, "wb") as f:
                    f.write(new_template.getbuffer())
                st.success(f"模板 '{template_name}' 保存成功!")

        # 模板列表
        templates = [f for f in os.listdir("templates") if f.endswith(".xlsx")]
        if templates:
            st.subheader("可用模板")
            selected_template = st.selectbox("选择模板", templates)
            return selected_template
        return None

# --------------------------------------------------------------------------
# 4. 主应用界面与逻辑
# --------------------------------------------------------------------------
def main():
    st.title("📊 智能Excel处理专家")
    st.markdown("""
    **使用说明:**
    1. 上传多个Excel数据文件
    2. 选择或上传模板文件
    3. 用自然语言描述处理需求
    4. 获取处理结果并下载
    """)
    st.markdown("---")

    # --- 初始化 ---
    if not os.path.exists("processed"):
        os.makedirs("processed")

    # --- 模板管理 ---
    selected_template = manage_templates()
    
    # --- API配置 ---
    api_key = st.sidebar.text_input(
        "DeepSeek API密钥",
        type="password",
        help="从DeepSeek官网获取API密钥"
    )
    
    if not api_key or not api_key.startswith("sk-"):
        st.warning("请输入有效的DeepSeek API密钥")
        st.stop()

    # --- 文件上传 ---
    uploaded_files = st.file_uploader(
        "上传一个或多个Excel文件",
        type=["xlsx", "xls"],
        accept_multiple_files=True
    )

    if not uploaded_files:
        st.info("请上传Excel文件以开始分析。")
        st.stop()
    
    # --- 数据预处理和预览 ---
    @st.cache_data(show_spinner="正在预处理数据...")
    def preprocess_data(uploaded_files):
        """统一数据预处理函数"""
        data_frames_dict = {}
        for file in uploaded_files:
            try:
                # 读取Excel文件并处理列名
                df = pd.read_excel(file)
                
                # 重命名未命名列
                df.columns = [f"Column_{i}" if col.startswith('Unnamed') else col 
                            for i, col in enumerate(df.columns, 1)]
                
                # 处理所有类型的空值
                df = df.fillna(0)  # 处理NaN
                df = df.replace(r'^\s*$', 0, regex=True)  # 处理空字符串和纯空格
                df = df.replace('', 0)  # 处理空字符串
                
                # 确保数值类型正确
                for col in df.columns:
                    if df[col].dtype == object:
                        try:
                            df[col] = pd.to_numeric(df[col], errors='ignore')
                        except:
                            pass
                
                data_frames_dict[file.name] = df
                
            except Exception as e:
                st.error(f"处理文件 {file.name} 时出错: {str(e)}")
                st.stop()
        
        return data_frames_dict

    # 执行预处理
    data_frames_dict = preprocess_data(uploaded_files)
    data_frames = list(data_frames_dict.values())
    
    # 显示预处理结果
    st.success(f"✅ 成功预处理 {len(data_frames)} 个文件")
    with st.expander("🔍 数据质量报告"):
        for name, df in data_frames_dict.items():
            st.write(f"**文件**: {name}")
            col1, col2 = st.columns(2)
            with col1:
                st.metric("总行数", len(df))
                st.metric("数值列", len(df.select_dtypes(include=['number']).columns))
            with col2:
                st.metric("总列数", len(df.columns))
                st.metric("文本列", len(df.select_dtypes(include=['object']).columns))
            st.dataframe(df.head(3))
    
    # --- 用户输入分析请求 ---
    st.subheader("📝 分析指令")
    
    # 示例指令选择
    example_instructions = {
        "基础分析": "计算各产品的销售总额并排序",
        "高级计算": "计算各产品的月环比增长率，并筛选增长率>10%的产品",
        "数据合并": "按客户ID合并两个表格，计算每个客户的总消费金额",
        "条件筛选": "筛选出销售额大于100万且利润率高于20%的产品",
        "时间序列": "按周汇总销售额，并计算4周移动平均值"
    }
    
    selected_example = st.selectbox("选择示例指令", list(example_instructions.keys()))
    analysis_prompt = st.text_area(
        "或自定义您的分析需求",
        height=100,
        value=example_instructions[selected_example],
        key="analysis_input"
    )
    
    # 高级选项
    with st.expander("⚙️ 高级计算设置"):
        precision = st.slider("计算精度(小数位数)", 0, 6, 2)
        use_advanced = st.checkbox("启用高级分析模式", help="支持更复杂的计算和自定义函数")

    # --- 执行分析按钮 ---
    if st.button("🚀 开始分析", type="primary", use_container_width=True):
        if not analysis_prompt.strip():
            st.warning("请输入您的分析需求。")
            st.stop()

        if not api_key or not api_key.startswith("sk-"):
            st.error("请输入一个有效的DeepSeek API密钥。")
            st.stop()

        with st.spinner("🧠 AI正在思考中，请稍候..."):
            try:
                # 初始化 LLM
                llm = DeepSeekLLM(api_key=api_key)

                # 初始化 PandasAI 的 SmartDatalake 并优化配置
                lake = SmartDatalake(
                    data_frames,
                    config={
                        "llm": llm,
                        "verbose": True,        # 在终端打印详细日志
                        "enable_cache": True,   # 启用缓存提高一致性
                        "max_retries": 3,       # 增加重试次数
                        "custom_instructions": """
                            你是一个严谨的数据分析师，请确保：
                            1. 计算结果精确到小数点后6位
                            2. 使用稳定的算法
                            3. 对结果进行双重验证
                            4. 避免随机性
                        """,
                        "response_parser": {
                            "structured": True,  # 强制结构化输出
                            "enforce_schema": True
                        }
                    }
                )
                
                # 优化分析请求提示词
                detailed_prompt = f"""
                请严格按照以下要求执行数据分析:
                
                分析需求: {analysis_prompt}
                
                计算要求:
                1. 使用精确计算方法，避免近似
                2. 分步计算并验证每一步结果
                3. 最终结果保留{precision}位小数
                4. 使用确定性算法，确保结果一致
                5. 对关键计算进行双重校验
                
                输出格式:
                1. 详细计算步骤及验证过程
                2. 中间结果表
                3. 最终汇总结果
                """
                
                # 响应处理器类
                class ResponseHandler:
                    @staticmethod
                    def handle_response(raw_response):
                        """安全处理各种类型的响应"""
                        st.write("🔍 开始处理响应数据...")
                        
                        # 类型安全检查
                        if raw_response is None:
                            raise ValueError("响应数据为空")
                            
                        # 调试信息
                        st.write(f"📊 响应类型: {type(raw_response)}")
                        if isinstance(raw_response, dict):
                            st.write(f"🔑 字典键: {list(raw_response.keys())}")
                        
                        # 分类型处理
                        handler_map = {
                            pd.DataFrame: ResponseHandler._handle_dataframe,
                            dict: ResponseHandler._handle_dict,
                            str: ResponseHandler._handle_string,
                            list: ResponseHandler._handle_list
                        }
                        
                        for data_type, handler in handler_map.items():
                            if isinstance(raw_response, data_type):
                                result = handler(raw_response)
                                if result is not None:
                                    return result
                                
                        # 未知类型处理
                        return ResponseHandler._handle_unknown(raw_response)
                    
                    @staticmethod
                    def _handle_dataframe(df):
                        """处理DataFrame类型响应"""
                        if not isinstance(df, pd.DataFrame):
                            return None
                            
                        st.write("✅ 获取到DataFrame响应")
                        if df.empty:
                            st.warning("⚠️ DataFrame为空")
                            return None
                            
                        st.write(f"📐 DataFrame形状: {df.shape}")
                        return df.copy()
                    
                    @staticmethod
                    def _handle_dict(data_dict):
                        """处理字典类型响应"""
                        if not isinstance(data_dict, dict):
                            return None
                            
                        st.write("✅ 获取到字典响应")
                        
                        # 尝试多种转换方式
                        converters = [
                            ResponseHandler._convert_dict_to_dataframe,
                            ResponseHandler._convert_nested_dict,
                            ResponseHandler._convert_simple_dict
                        ]
                        
                        for converter in converters:
                            try:
                                result = converter(data_dict)
                                if result is not None and not result.empty:
                                    return result
                            except Exception as e:
                                st.warning(f"⚠️ 字典转换尝试失败: {str(e)}")
                                
                        return None
                    
                    @staticmethod
                    def _handle_string(text):
                        """处理字符串类型响应"""
                        if not isinstance(text, str) or not text.strip():
                            return None
                            
                        st.write("✅ 获取到文本响应")
                        return pd.DataFrame({"结果": [text.strip()]})
                    
                    @staticmethod
                    def _handle_list(data_list):
                        """处理列表类型响应"""
                        if not isinstance(data_list, list):
                            return None
                            
                        st.write("✅ 获取到列表响应")
                        try:
                            return pd.DataFrame(data_list)
                        except Exception as e:
                            st.warning(f"⚠️ 列表转换失败: {str(e)}")
                            return None
                    
                    @staticmethod
                    def _handle_unknown(data):
                        """处理未知类型响应"""
                        st.warning("⚠️ 未知响应类型")
                        try:
                            return pd.DataFrame([str(data)])
                        except Exception as e:
                            st.error(f"❌ 无法处理响应数据: {str(e)}")
                            return None
                    
                    @staticmethod
                    def _convert_dict_to_dataframe(data_dict):
                        """标准字典转换"""
                        return pd.DataFrame(data_dict)
                    
                    @staticmethod
                    def _convert_nested_dict(data_dict):
                        """嵌套字典转换"""
                        if not all(isinstance(v, (dict, list, pd.Series)) for v in data_dict.values()):
                            return None
                        return pd.DataFrame.from_dict(data_dict, orient='columns')
                    
                    @staticmethod
                    def _convert_simple_dict(data_dict):
                        """简单键值对转换"""
                        return pd.DataFrame([data_dict])
                
                # 执行分析并处理响应
                response = None
                for attempt in range(3):
                    st.write(f"🔄 分析尝试 {attempt + 1}/3")
                    try:
                        raw_response = lake.chat(detailed_prompt)
                        response = ResponseHandler.handle_response(raw_response)
                        
                        if response is not None and not response.empty:
                            st.success("✅ 成功获取有效响应")
                            break
                            
                    except Exception as e:
                        st.error(f"⚠️ 分析失败: {str(e)}")
                        st.exception(e)
                        time.sleep(1)
                
                if response is None:
                    raise ValueError("❌ 所有分析尝试均失败，请检查输入数据")
                elif response.empty:
                    raise ValueError("⚠️ 获取到空结果，请调整分析参数")
                
                st.subheader("🧮 详细计算过程")
                st.markdown("---")
                
                # --- 智能结果展示 ---
                if response is None:
                    st.warning("分析未能返回有效结果，请尝试调整您的问题。")
                
                elif isinstance(response, pd.DataFrame):
                    try:
                        # 安全处理响应数据
                        st.write("?? 正在准备响应数据...")
                        
                        # 确保响应是DataFrame
                        response_df = response.copy() if isinstance(response, pd.DataFrame) else pd.DataFrame([response])
                        
                        # 结果展示处理器
                        class ResultDisplayer:
                            @staticmethod
                            def show_calculation_steps(df):
                                if "计算步骤" not in df.columns:
                                    return
                                    
                                with st.expander("📝 计算步骤详解", expanded=True):
                                    steps = df["计算步骤"]
                                    if isinstance(steps, pd.Series) and len(steps) > 0:
                                        st.write("### 详细计算流程")
                                        for i, step in enumerate(steps.dropna(), 1):
                                            st.write(f"{i}. {step}")
                                    else:
                                        st.info("ℹ️ 无详细计算步骤记录")
                            
                            @staticmethod
                            def show_interim_results(df, precision):
                                if "中间结果" not in df.columns:
                                    return
                                    
                                with st.expander("🔍 查看中间结果", expanded=False):
                                    interim = df["中间结果"]
                                    if isinstance(interim, pd.DataFrame) and not interim.empty:
                                        st.write("### 中间计算结果")
                                        st.dataframe(interim.style.format(precision=precision))
                                    elif isinstance(interim, (pd.Series, list)) and len(interim) > 0:
                                        st.write("### 中间计算结果")
                                        st.write(interim)
                                    else:
                                        st.info("ℹ️ 无中间结果记录")
                        
                        # 显示计算结果
                        ResultDisplayer.show_calculation_steps(response_df)
                        ResultDisplayer.show_interim_results(response_df, precision)
                        
                        # 安全准备最终结果
                        final_result = response_df.copy()
                        cols_to_drop = [col for col in ["计算步骤", "中间结果"] 
                                      if col in response_df.columns]
                        if cols_to_drop:
                            final_result = final_result.drop(columns=cols_to_drop, errors='ignore')
                        
                        # 显示最终结果
                        st.subheader("💡 最终分析结果")
                        if isinstance(final_result, pd.DataFrame):
                            if not final_result.empty:
                                st.dataframe(final_result)
                            else:
                                st.warning("结果数据为空")
                        elif isinstance(final_result, str):
                            st.write(final_result)
                        else:
                            st.warning(f"无法识别的结果格式: {type(final_result)}")
                            st.write(str(final_result))
                            
                    except Exception as e:
                        st.error(f"处理分析结果时出错: {str(e)}")
                        st.info("正在尝试显示原始结果...")
                        st.dataframe(response)
                    
                    # 结果回填选项 - 适配新格式
                    st.markdown("---")
                    st.markdown("### 📤 结果输出方式")
                    with st.container():
                        col1, col2 = st.columns([3,7])
                        with col1:
                            fill_original = st.checkbox(
                                "🔁 回填到原文件", 
                                value=True,
                                help="将分析结果回填到原始Excel文件，保留所有格式和样式"
                            )
                        with col2:
                            if fill_original:
                                st.success("已启用回填功能 - 最终结果将保存回原始文件")
                            else:
                                st.info("将生成包含完整计算过程的新文件")
                    
                    # 添加完整报告下载选项
                    st.download_button(
                        label="📥 下载完整计算报告",
                        data=response.to_csv(index=False).encode("utf-8"),
                        file_name="detailed_analysis_report.csv",
                        mime="text/csv",
                        help="下载包含所有计算步骤和中间结果的完整报告"
                    )
                    st.markdown("---")
                    
                    if fill_original:
                        from openpyxl import load_workbook
                        from io import BytesIO
                        
                        with st.status("正在处理回填操作...", expanded=True) as status:
                            # 处理每个原始文件
                            for file_idx, file in enumerate(uploaded_files):
                                try:
                                    st.write(f"🔧 正在处理文件: {file.name}...")
                                    
                                    # 读取原始文件
                                    file.seek(0)
                                    wb = load_workbook(file)
                                    ws = wb.active
                                    
                                    # 获取原文件列名和格式
                                    original_columns = []
                                    column_formats = {}
                                    for col in range(1, ws.max_column+1):
                                        col_name = ws.cell(row=1, column=col).value
                                        original_columns.append(col_name)
                                        # 记录列格式(从第二行获取)
                                        if ws.max_row >= 2:
                                            sample_cell = ws.cell(row=2, column=col)
                                            column_formats[col_name] = {
                                                'number_format': sample_cell.number_format,
                                                'alignment': sample_cell.alignment,
                                                'font': sample_cell.font,
                                                'fill': sample_cell.fill,
                                                'border': sample_cell.border
                                            }
                                    
                                    # 安全地准备结果数据
                                    result_df = response.copy() if isinstance(response, pd.DataFrame) else pd.DataFrame()
                                    
                                    # 列名匹配检查和处理
                                    original_columns = [col for col in original_columns 
                                                      if col is not None and not str(col).startswith('Unnamed')]
                                    result_columns = [col for col in result_df.columns 
                                                    if col is not None and not str(col).startswith('Unnamed')]
                                    
                                    # 安全地检查列名匹配
                                    missing_cols = [col for col in original_columns 
                                                  if col not in result_columns and pd.notna(col)]
                                    extra_cols = [col for col in result_columns 
                                                if col not in original_columns and pd.notna(col)]
                                    
                                    if missing_cols:
                                        st.warning(f"⚠️ 原文件中有 {len(missing_cols)} 列在结果中不存在，将保留为空列")
                                        st.write("缺失列:", missing_cols)
                                        st.info("提示：未命名列(Unnamed)已自动忽略")
                                    
                                    if extra_cols:
                                        st.warning(f"⚠️ 结果中有 {len(extra_cols)} 列在原文件中不存在，将不会被回填")
                                        st.write("额外列:", extra_cols)
                                        st.info("提示：未命名列(Unnamed)已自动忽略")
                                    
                                    # 确保保留所有原始列
                                    for col in original_columns:
                                        if col not in result_df.columns:
                                            result_df[col] = None
                                    
                                    # 重新排列结果列以匹配原文件顺序
                                    result_df = result_df[original_columns]
                                    
                                    # 写入数据
                                    st.write("📝 正在回填数据...")
                                    for row_idx, row_data in enumerate(result_df.itertuples(index=False), start=2):
                                        for col_idx, value in enumerate(row_data, start=1):
                                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                                            # 应用原格式
                                            col_name = original_columns[col_idx-1]
                                            if col_name in column_formats:
                                                fmt = column_formats[col_name]
                                                cell.number_format = fmt['number_format']
                                                cell.alignment = fmt['alignment']
                                                cell.font = fmt['font']
                                                cell.fill = fmt['fill']
                                                cell.border = fmt['border']
                                    
                                    # 保存文件
                                    st.write("💾 正在保存文件...")
                                    output = BytesIO()
                                    wb.save(output)
                                    output.seek(0)
                                    
                                    status.update(label=f"✅ 文件 {file.name} 回填完成!", state="complete")
                                    
                                    # 提供下载
                                    st.download_button(
                                        label=f"📥 下载回填后的文件: {file.name}",
                                        data=output,
                                        file_name=f"updated_{file.name}",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        help="点击下载已回填结果的Excel文件，将保留原文件的所有格式和公式"
                                    )
                                    
                                    # 显示回填摘要
                                    with st.expander(f"🔍 回填摘要: {file.name}"):
                                        st.write(f"原文件列数: {len(original_columns)}")
                                        st.write(f"结果数据列数: {len(response.columns)}")
                                        st.write(f"匹配列数: {len(set(original_columns) & set(response.columns))}")
                                        st.dataframe(result_df.head(3))
                                    
                                except Exception as e:
                                    st.error(f"❌ 回填文件 {file.name} 时出错: {str(e)}")
                                    with st.expander("🛠️ 详细调试信息"):
                                        st.write("原文件列名:", original_columns)
                                        st.write("结果列名:", response.columns.tolist())
                                        st.write("错误详情:", str(e))
                                    
                                    st.error("""
                                    🚨 常见解决方法:
                                    1. 检查列名是否完全匹配
                                    2. 修改分析指令明确指定列名
                                    3. 检查文件是否受保护
                                    4. 尝试简化分析需求
                                    """)
                    else:
                        # 提供CSV下载
                        csv = response.to_csv(index=False).encode("utf-8-sig")
                        st.download_button(
                            label="📥 下载结果 (CSV)",
                            data=csv,
                            file_name="analysis_result.csv",
                            mime="text/csv",
                        )
                
                elif isinstance(response, (str, int, float)):
                    st.markdown(f"### {response}")

                elif isinstance(response, dict) and response.get("type") == "plot":
                    st.image(response["value"], caption="分析图表")
                
                else:
                    # 对于其他未知类型的返回结果，直接以文本形式输出
                    st.text("未能识别的返回类型，原始输出如下：")
                    st.code(str(response))
                    
            except Exception as e:
                st.error(f"分析过程中发生严重错误：\n\n{e}")

# --------------------------------------------------------------------------
# 4. 启动应用
# --------------------------------------------------------------------------
if __name__ == "__main__":
    main()